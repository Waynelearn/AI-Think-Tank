"""Process uploaded files into text that can be fed to agents as context."""

import csv
import io
import os
import html.parser


def process_file(filename: str, content: bytes) -> str:
    """Route a file to the appropriate processor based on extension. Returns extracted text."""
    ext = os.path.splitext(filename)[1].lower()
    processors = {
        ".csv": _process_csv,
        ".xlsx": _process_excel,
        ".xls": _process_excel,
        ".pdf": _process_pdf,
        ".html": _process_html,
        ".htm": _process_html,
        ".txt": _process_text,
        ".md": _process_text,
        ".json": _process_text,
        ".log": _process_text,
        ".py": _process_text,
        ".js": _process_text,
        ".ts": _process_text,
        ".docx": _process_docx,
        ".png": _process_image,
        ".jpg": _process_image,
        ".jpeg": _process_image,
        ".gif": _process_image,
        ".webp": _process_image,
        ".mp4": _process_video,
        ".mov": _process_video,
        ".avi": _process_video,
        ".mkv": _process_video,
    }
    processor = processors.get(ext)
    if not processor:
        return f"[Unsupported file type: {ext}. File name: {filename}]"
    try:
        return processor(filename, content)
    except Exception as e:
        return f"[Error processing {filename}: {e}]"


def _process_csv(filename: str, content: bytes) -> str:
    text = content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return "[Empty CSV file]"
    # Show header + up to 50 data rows
    header = rows[0]
    data_rows = rows[1:51]
    lines = [f"CSV file: {filename} ({len(rows)-1} data rows)"]
    lines.append("Columns: " + " | ".join(header))
    lines.append("---")
    for row in data_rows:
        lines.append(" | ".join(row))
    if len(rows) > 51:
        lines.append(f"... ({len(rows)-51} more rows truncated)")
    return "\n".join(lines)


def _process_excel(filename: str, content: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    lines = [f"Excel file: {filename}"]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"\n## Sheet: {sheet_name}")
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            if row_count >= 50:
                lines.append("... (rows truncated)")
                break
            lines.append(" | ".join(str(c) if c is not None else "" for c in row))
            row_count += 1
    wb.close()
    return "\n".join(lines)


def _process_pdf(filename: str, content: bytes) -> str:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(content))
    lines = [f"PDF file: {filename} ({len(reader.pages)} pages)"]
    for i, page in enumerate(reader.pages[:30]):  # cap at 30 pages
        text = page.extract_text()
        if text and text.strip():
            lines.append(f"\n--- Page {i+1} ---\n{text.strip()}")
    if len(reader.pages) > 30:
        lines.append(f"\n... ({len(reader.pages) - 30} more pages truncated)")
    return "\n".join(lines)


def _process_html(filename: str, content: bytes) -> str:
    text = content.decode("utf-8", errors="replace")

    class _HTMLTextExtractor(html.parser.HTMLParser):
        def __init__(self):
            super().__init__()
            self.parts = []

        def handle_data(self, data):
            self.parts.append(data)

    extractor = _HTMLTextExtractor()
    extractor.feed(text)
    extracted = " ".join(extractor.parts).strip()
    if len(extracted) > 10000:
        extracted = extracted[:10000] + "\n... (truncated)"
    return f"HTML file: {filename}\n\n{extracted}"


def _process_text(filename: str, content: bytes) -> str:
    text = content.decode("utf-8", errors="replace")
    if len(text) > 10000:
        text = text[:10000] + "\n... (truncated)"
    return f"File: {filename}\n\n{text}"


def _process_docx(filename: str, content: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(content))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    text = "\n".join(paragraphs)
    if len(text) > 10000:
        text = text[:10000] + "\n... (truncated)"
    return f"Word document: {filename}\n\n{text}"


def _process_image(filename: str, content: bytes) -> str:
    from PIL import Image
    img = Image.open(io.BytesIO(content))
    info = f"Image file: {filename} ({img.format}, {img.size[0]}x{img.size[1]}, mode={img.mode})"
    return f"[{info} — image content cannot be read as text, but the file has been attached for reference]"


def _process_video(filename: str, content: bytes) -> str:
    size_mb = len(content) / (1024 * 1024)
    return f"[Video file: {filename} ({size_mb:.1f} MB) — video content noted for discussion context]"
